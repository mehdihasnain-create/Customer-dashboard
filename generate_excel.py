"""
Klarity Support — Excel Report Generator
=========================================
Fetches all actionable tickets from Zendesk view 25994134990364 (the curated
"All Actionable" view) plus the channel views, applies the same Python-side
filtering used in the Streamlit dashboard, and writes a fully formatted Excel
workbook with pivot-style summary tables and embedded charts.

Usage:
    python3 generate_excel.py \
        --email you@klaritylaw.com \
        --token YOUR_API_TOKEN \
        --weeks 4                     # number of recent complete weeks to cover
        --output klarity_report.xlsx  # optional, defaults to klarity_YYYYMMDD.xlsx

Requirements:
    pip install requests openpyxl
"""

import argparse
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta

import requests
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ── CONSTANTS (mirrors klarity_dashboard.py) ──────────────────────────────────
SUBDOMAIN = "klarity6695"

FILTERS = {
    "Actionable":      "25994134990364",   # ← source of truth for all actionable tickets
    "AI Bot":          "25251640968348",
    "Messaging/Chat":  "23607369574812",
    "Failed Ops":      "17237919534108",   # exclusion view
    "Architect":       "24101684048412",
    "High Priority":   "23645136614684",
}

KLARITY_DOMAINS = ("@klaritylaw.com", "@klarity.ai")
ARCHITECT_EMAIL = "architect@klarity.ai"

KNOWN_CUSTOMERS = [
    "zuora", "mongodb", "stripe", "doordash", "sentinelone", "miro",
    "linkedin", "aven", "quench", "kimball", "ipw", "uhg", "ramp",
    "cloudflare", "brex", "rippling", "workday", "salesforce", "hubspot",
]
CUSTOMER_DISPLAY = {
    "aven": "Aven Hospitality", "doordash": "DoorDash",
    "sentinelone": "SentinelOne", "linkedin": "LinkedIn",
    "mongodb": "MongoDB", "hubspot": "HubSpot",
    "cloudflare": "Cloudflare", "ipw": "IPW", "uhg": "UHG",
    "ramp": "Ramp", "miro": "Miro", "zuora": "Zuora",
    "stripe": "Stripe", "quench": "Quench", "kimball": "Kimball",
    "brex": "Brex", "rippling": "Rippling", "workday": "Workday",
    "salesforce": "Salesforce",
}

CATEGORIES = [
    ("Architect Run Failure / Error running operation",
     re.compile(r"error running|error while running|run fail|architect run|flow run|operation.*fail", re.I),
     ["architect_run_failure", "operation___workflow_fail", "bugs/issues"]),
    ("Unable to login / Access issues",
     re.compile(r"unable.*(login|log in|sign in)|login issue|access|password|temp.*pass|sso|new.*workspace|workspace.*modif", re.I),
     ["login", "access", "workspace"]),
    ("Table matching / Data issues",
     re.compile(r"mismatch|match|table|missing.*line|incorrect|revenue|duplicate|deal|renewal|discrepan", re.I),
     ["matching", "table_matching", "revenue", "integration_issue"]),
    ("AI / Transcript / Hallucination issues",
     re.compile(r"hallucin|transcript|ai interviewer|coach|indexing|insight|pilot", re.I),
     ["transcript_processing/hallucinations", "indexing/insights_errors", "coach"]),
    ("Screenshot / Image / SOP issues",
     re.compile(r"screenshot|image placement|sop|screen share", re.I), []),
    ("Token / Time limits",
     re.compile(r"token|time limit|update with ai|anthropic token", re.I), []),
    ("Timeouts / Performance",
     re.compile(r"timeout|concurrent|performance|slow", re.I), []),
    ("Feedback / Feature Request",
     re.compile(r"feedback|feature request", re.I), ["feedback"]),
    ("Others", None, []),
]

TOP_ISSUES = [
    ("Run Failure / Architect errors",  re.compile(r"error running|error while running|run fail|flow run", re.I)),
    ("Login / Access / Workspace",      re.compile(r"login|access|password|workspace|temp.*pass", re.I)),
    ("Table matching / Data issues",    re.compile(r"mismatch|match|table|missing|duplicate|deal|revenue|discrepan", re.I)),
    ("AI / Transcript / Hallucination", re.compile(r"hallucin|transcript|ai interviewer|coach|pilot", re.I)),
]

# ── COLOURS ───────────────────────────────────────────────────────────────────
ORANGE     = "E8612C"
DARK       = "1A1A1A"
LIGHT_BG   = "F7F5F2"
WHITE      = "FFFFFF"
GREY       = "888888"
LIGHT_GREY = "EEEEEE"
ROW_ALT    = "FAFAFA"
EXCL_BG    = "FDF0EA"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr_font(bold=True, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def body_font(bold=False, color=DARK, size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def thin_border():
    s = Side(border_style="thin", color=LIGHT_GREY)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── WEEK HELPERS ──────────────────────────────────────────────────────────────
def monday_of(d):
    return d - timedelta(days=d.weekday())

def iso_week_bounds(year, week_num):
    jan4      = date(year, 1, 4)
    week1_mon = monday_of(jan4)
    start     = week1_mon + timedelta(weeks=week_num - 1)
    end       = start + timedelta(days=6)
    return start, end

def since_n_weeks(n):
    today       = date.today()
    this_monday = monday_of(today)
    return (this_monday - timedelta(weeks=n)).isoformat()

def in_week(t, w):
    c = (t.get("created_at") or "")[:10]
    return w["start"] <= c < w["end"]

def get_recent_weeks(n):
    today       = date.today()
    this_monday = monday_of(today)
    weeks = []
    for i in range(n - 1, -1, -1):
        start = this_monday - timedelta(weeks=i + 1)
        end   = start + timedelta(days=7)
        wn    = start.isocalendar()[1]
        weeks.append({
            "label":   f"Wk {wn}",
            "week_num": wn,
            "start":   start.isoformat(),
            "end":     end.isoformat(),
            "display": f"{start.strftime('%b %d')} – {(end - timedelta(days=1)).strftime('%b %d')}",
        })
    return weeks

def get_all_year_weeks():
    today    = date.today()
    cur_year = today.isocalendar()[0]
    cur_week = today.isocalendar()[1]
    weeks    = []
    for wn in range(1, cur_week + 1):
        start, end = iso_week_bounds(cur_year, wn)
        weeks.append({
            "label":    f"Wk {wn}",
            "week_num": wn,
            "start":    start.isoformat(),
            "end":      (end + timedelta(days=1)).isoformat(),
            "display":  f"{start.strftime('%b %d')} - {end.strftime('%b %d')}",
        })
    return weeks

# ── TICKET CLASSIFIERS ────────────────────────────────────────────────────────
def tags_of(t):
    return set(t.get("tags") or [])

def is_backend_alert(t):
    return bool(re.search(r"\[BACKEND ALERT\]", t.get("subject", "") or "", re.I))

def is_internal_teams(t):
    return "internal_teams" in tags_of(t)

def is_klarity_staff(email):
    return any(email.lower().endswith(d) for d in KLARITY_DOMAINS)

def is_internal_run_failure(t):
    subj = t.get("subject", "") or ""
    if not re.search(r"error running operation|error while running", subj, re.I):
        return False
    req_email = (t.get("_requester_email") or "").lower()
    if req_email == ARCHITECT_EMAIL:
        return False
    return "automated_architect" in tags_of(t)

def is_excluded(t, fo_ids):
    if t["id"] in fo_ids:          return True
    if t.get("status") == "new":   return True
    if is_backend_alert(t):        return True
    if is_internal_teams(t):       return True
    if is_internal_run_failure(t): return True
    return False

def is_open(t):   return t.get("status") in ("open", "pending")
def is_closed(t): return t.get("status") in ("solved", "closed")

def res_hours(t):
    if not is_closed(t): return None
    try:
        c  = datetime.fromisoformat(t["created_at"].replace("Z", "+00:00"))
        s  = t.get("solved_at")
        if not s: return None
        sd = datetime.fromisoformat(s.replace("Z", "+00:00"))
        h  = round((sd - c).total_seconds() / 3600)
        return h if h >= 0 else None
    except Exception:
        return None

def med(v): return sorted(v)[len(v) // 2] if v else None
def avg(v): return round(sum(v) / len(v)) if v else None
def p90(v): return sorted(v)[int(len(v) * 0.9)] if v else None

def categorize(tickets):
    result = {label: [] for label, _, __ in CATEGORIES}
    for t in tickets:
        subj = t.get("subject", "") or ""
        tgs  = tags_of(t)
        for label, rx, tag_list in CATEGORIES:
            if rx is None or rx.search(subj) or any(tg in tgs for tg in tag_list):
                result[label].append(t)
                break
    return result

def extract_customer(t):
    subj = t.get("subject", "") or ""
    tgs  = tags_of(t)
    m = re.search(r"<>([^|<>]+)\s*\|\|", subj)
    if m:
        return m.group(1).strip()
    if "||" in subj:
        part = subj.split("||")[0].strip()
        part = re.sub(r"^\[[^\]]*\]\s*", "", part).strip()
        if 1 < len(part) < 50 and part.lower() not in ("klarity", "klarity ai", ""):
            return part
    org = (t.get("_org_name") or "").strip()
    if org and org.lower() not in ("klarity", "klarity ai", "klarity law", ""):
        return org
    for k in KNOWN_CUSTOMERS:
        if any(k.lower() in tg.lower() for tg in tgs):
            return CUSTOMER_DISPLAY.get(k, k.title())
    subj_lower = subj.lower()
    for k in KNOWN_CUSTOMERS:
        if k.lower() in subj_lower:
            return CUSTOMER_DISPLAY.get(k, k.title())
    return "Other"

def arch_bucket(t):
    req = (t.get("_requester_email") or "").lower()
    if req == ARCHITECT_EMAIL:
        return "failure_notification"
    if is_klarity_staff(req):
        return "internal"
    return "customer"

def ticket_channel(t, ch_id_sets):
    """Return the channel name for a ticket based on view membership."""
    tid = t["id"]
    for name, id_set in ch_id_sets.items():
        if tid in id_set:
            return name
    return "Other"

# ── ZENDESK API ───────────────────────────────────────────────────────────────
def zd_get(email, token, path, params=None):
    r = requests.get(
        f"https://{SUBDOMAIN}.zendesk.com/api/v2{path}",
        auth=(f"{email}/token", token), params=params, timeout=30
    )
    r.raise_for_status()
    return r.json()

def fetch_view(email, token, view_id, since=None):
    """Fetch all tickets from a Zendesk view with enrichment sideloads."""
    tickets, page = [], 1
    while True:
        try:
            data = zd_get(
                email, token,
                f"/views/{view_id}/tickets.json",
                {"per_page": 100, "page": page,
                 "include": "users,organizations,metric_sets"},
            )
            users   = {u["id"]: u.get("email", "") for u in data.get("users", [])}
            orgs    = {o["id"]: o.get("name", "")  for o in data.get("organizations", [])}
            metrics = {m["ticket_id"]: m            for m in data.get("metric_sets", [])}
            batch   = data.get("tickets", [])
            raw_count = len(batch)

            for t in batch:
                t["_requester_email"] = users.get(t.get("requester_id"), "")
                t["_org_name"]        = orgs.get(t.get("organization_id"), "")
                ms = metrics.get(t["id"])
                if ms:
                    t["solved_at"] = ms.get("solved_at")

            if since:
                batch = [t for t in batch if (t.get("created_at") or "")[:10] >= since]

            tickets.extend(batch)
            if not data.get("next_page") or raw_count < 100:
                break
            page += 1
        except Exception as e:
            print(f"  Warning: view {view_id} page {page} error: {e}")
            break
    return tickets

# ── DATA LOADING ──────────────────────────────────────────────────────────────
def load_all_data(email, token, since):
    print("  [1/6] Fetching Failed Ops view (exclusion list)…")
    fo_all   = fetch_view(email, token, FILTERS["Failed Ops"])
    fo_ids   = {t["id"] for t in fo_all}

    print("  [2/6] Fetching Actionable view (all actionable tickets)…")
    actionable = fetch_view(email, token, FILTERS["Actionable"])

    print("  [3/6] Fetching channel views…")
    ch_raw = {}
    for name in ["AI Bot", "Messaging/Chat", "Architect", "High Priority"]:
        ch_raw[name] = fetch_view(email, token, FILTERS[name], since=since)

    # Build channel ID sets for channel attribution
    ch_id_sets = {name: {t["id"] for t in tickets}
                  for name, tickets in ch_raw.items()}

    print("  [4/6] Enriching actionable tickets…")
    # Enrich actionable tickets with email/org/solved_at from channel views
    all_side = [t for vl in ch_raw.values() for t in vl] + fo_all
    email_map = {t["id"]: t["_requester_email"] for t in all_side if t.get("_requester_email")}
    org_map   = {t["id"]: t["_org_name"]        for t in all_side if t.get("_org_name")}
    solve_map = {t["id"]: t.get("solved_at")    for t in all_side if t.get("solved_at")}

    for t in actionable:
        if not t.get("_requester_email"):
            t["_requester_email"] = email_map.get(t["id"], "")
        if not t.get("_org_name"):
            t["_org_name"]        = org_map.get(t["id"], "")
        if not t.get("solved_at"):
            t["solved_at"]        = solve_map.get(t["id"])

    print("  [5/6] Applying exclusion rules…")
    # Actionable view is already curated — apply extra Python-side safety checks
    real = [t for t in actionable if not is_excluded(t, fo_ids)]

    # Date-filter for selected range
    ranged   = [t for t in real if (t.get("created_at") or "")[:10] >= since]
    open_t   = [t for t in ranged if is_open(t)]
    closed_t = [t for t in ranged if is_closed(t)]

    # All-open (no date filter) for "Unsolved in Group"
    unsolved = [t for t in real if is_open(t)]

    # Channel-ranged (within date range)
    ch = {}
    for name in ["AI Bot", "Messaging/Chat", "Architect", "High Priority"]:
        t_in_range = [t for t in ch_raw[name]
                      if not is_excluded(t, fo_ids)
                      and (t.get("created_at") or "")[:10] >= since]
        ch[name] = {
            "tickets": t_in_range,
            "open":    [t for t in t_in_range if is_open(t)],
            "closed":  [t for t in t_in_range if is_closed(t)],
        }

    # Architect sub-buckets
    arch_all = ch["Architect"]["tickets"]
    arch_b   = {k: {"tickets": [], "open": [], "closed": []}
                for k in ("customer", "internal", "failure_notification")}
    for t in arch_all:
        k = arch_bucket(t)
        arch_b[k]["tickets"].append(t)
        if is_open(t):   arch_b[k]["open"].append(t)
        if is_closed(t): arch_b[k]["closed"].append(t)

    print("  [6/6] Computing derived metrics…")
    # Category performance (last 4 complete weeks — fixed window)
    s4      = since_n_weeks(4)
    real_4w = [t for t in real if (t.get("created_at") or "")[:10] >= s4]
    cat_map  = categorize(real_4w)
    cat_perf = []
    for label, _, __ in CATEGORIES:
        tix   = cat_map[label]
        if not tix: continue
        times = [h for t in tix if (h := res_hours(t)) is not None]
        cat_perf.append({"label": label, "count": len(tix),
                         "median": med(times), "avg": avg(times), "p90": p90(times)})
    cat_perf.sort(key=lambda x: -x["count"])

    # Week-by-week trend
    all_weeks = get_all_year_weeks()
    for w in all_weeks:
        w["tickets"] = [t for t in real if in_week(t, w)]
        w["count"]   = len(w["tickets"])

    # Customer breakdown (ranged)
    cust_map = defaultdict(lambda: {"open": 0, "solved": 0})
    for t in ranged:
        cust = extract_customer(t)
        if is_open(t):   cust_map[cust]["open"]   += 1
        else:            cust_map[cust]["solved"]  += 1

    # Daily volume (ranged)
    daily = Counter((t.get("created_at") or "")[:10] for t in ranged
                    if t.get("created_at"))

    # Top issues over the last 4 shown weeks
    recent_weeks = get_recent_weeks(min(4, len(all_weeks)))
    for w in recent_weeks:
        w["tickets"] = [t for t in real if in_week(t, w)]

    # Status distribution (ranged)
    status_counts = Counter(t.get("status", "unknown").capitalize() for t in ranged)

    # Resolution time by category (ranged)
    cat_map_range = categorize(ranged)
    cat_range_stats = []
    for label, _, __ in CATEGORIES:
        tix   = cat_map_range[label]
        times = [h for t in tix if (h := res_hours(t)) is not None]
        cat_range_stats.append({
            "label": label, "count": len(tix),
            "median": med(times), "avg": avg(times), "p90": p90(times)
        })

    return dict(
        real=real, ranged=ranged, open_t=open_t, closed_t=closed_t,
        unsolved=unsolved, fo_count=len(fo_all),
        ch=ch, arch_b=arch_b,
        cat_perf=cat_perf, all_weeks=all_weeks,
        recent_weeks=recent_weeks, cust_map=cust_map,
        daily=daily, status_counts=status_counts,
        cat_range_stats=cat_range_stats, since=since,
    )

# ── EXCEL HELPERS ─────────────────────────────────────────────────────────────
def write_header_row(ws, row, cols, col_start=1):
    """Write a styled orange header row."""
    for i, col in enumerate(cols):
        c = ws.cell(row=row, column=col_start + i, value=col)
        c.fill      = fill(ORANGE)
        c.font      = hdr_font()
        c.alignment = center()
        c.border    = thin_border()

def write_data_row(ws, row, values, col_start=1, bold=False, bg=WHITE):
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=col_start + i, value=val)
        c.font      = body_font(bold=bold)
        c.alignment = left() if i == 0 else center()
        c.fill      = fill(bg)
        c.border    = thin_border()

def write_total_row(ws, row, values, col_start=1):
    write_data_row(ws, row, values, col_start=col_start, bold=True, bg=LIGHT_BG)
    ws.cell(row=row, column=col_start).font = Font(bold=True, color=DARK, size=10, name="Calibri")

def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def section_title(ws, row, title, col_span, col_start=1):
    c = ws.cell(row=row, column=col_start, value=title)
    c.font      = Font(bold=True, size=13, color=DARK, name="Calibri")
    c.fill      = fill(LIGHT_BG)
    c.alignment = left()
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row,   end_column=col_start + col_span - 1
    )

def add_ws(wb, title):
    ws = wb.create_sheet(title=title)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ORANGE
    return ws

def fmt_h(v):
    return f"{v}h" if v is not None else "—"

# ── SHEET WRITERS ─────────────────────────────────────────────────────────────

def write_summary(wb, data, since, weeks_n):
    ws = add_ws(wb, "Summary")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    # Title
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value     = "Klarity Support — Summary"
    c.font      = Font(bold=True, size=16, color=ORANGE, name="Calibri")
    c.alignment = left()

    ws.merge_cells("A2:B2")
    sub = ws["A2"]
    sub.value     = f"Period: {since}  →  {date.today().isoformat()}   |   Generated {datetime.now().strftime('%d %b %Y %H:%M')}"
    sub.font      = Font(italic=True, size=10, color=GREY, name="Calibri")
    sub.alignment = left()

    metrics = [
        ("Tickets Received (period)",    len(data["ranged"])),
        ("Open",                          len(data["open_t"])),
        ("Solved / Closed",               len(data["closed_t"])),
        ("Resolution Rate",
         f"{round(len(data['closed_t']) / len(data['ranged']) * 100)}%"
         if data["ranged"] else "—"),
        ("Unsolved in Group (live)",      len(data["unsolved"])),
        ("High Priority (period)",        len(data["ch"]["High Priority"]["tickets"])),
        ("Failed Ops (excluded)",         data["fo_count"]),
        ("Total Actionable (all time)",   len(data["real"])),
    ]

    write_header_row(ws, 4, ["Metric", "Value"])
    for i, (metric, val) in enumerate(metrics):
        bg = WHITE if i % 2 == 0 else ROW_ALT
        write_data_row(ws, 5 + i, [metric, val], bg=bg)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[4].height = 20


def write_raw_data(wb, data):
    ws = add_ws(wb, "Raw Tickets")

    headers = [
        "Ticket ID", "Subject", "Status", "Channel", "Category",
        "Customer", "Requester Email", "Created Date", "Week",
        "Resolution Hours", "Arch Bucket"
    ]
    write_header_row(ws, 1, headers)

    # Pre-compute categories
    cat_map = categorize(data["ranged"])
    ticket_cat = {}
    for label, _, __ in CATEGORIES:
        for t in cat_map[label]:
            ticket_cat[t["id"]] = label

    # Build channel membership
    ch_id = {}
    for name, ch in data["ch"].items():
        for t in ch["tickets"]:
            ch_id.setdefault(t["id"], name)

    for i, t in enumerate(
        sorted(data["ranged"], key=lambda x: x.get("created_at", ""), reverse=True)
    ):
        wn = ""
        ca = (t.get("created_at") or "")[:10]
        if ca:
            try:
                d  = date.fromisoformat(ca)
                wn = f"Wk {d.isocalendar()[1]}"
            except Exception:
                pass

        ab = arch_bucket(t) if ch_id.get(t["id"]) == "Architect" else ""

        row_vals = [
            t["id"],
            (t.get("subject") or "").strip()[:120],
            (t.get("status") or "").capitalize(),
            ch_id.get(t["id"], "Other"),
            ticket_cat.get(t["id"], "Others"),
            extract_customer(t),
            t.get("_requester_email") or "",
            ca,
            wn,
            res_hours(t),
            ab,
        ]
        bg = WHITE if i % 2 == 0 else ROW_ALT
        write_data_row(ws, 2 + i, row_vals, bg=bg)

    # Column widths
    widths = {"A": 12, "B": 55, "C": 10, "D": 18, "E": 38, "F": 22,
              "G": 28, "H": 13, "I": 8, "J": 18, "K": 22}
    set_col_widths(ws, widths)

    # Table style (enables Excel auto-filter)
    from openpyxl.worksheet.table import Table, TableStyleInfo
    tbl = Table(
        displayName="TicketData",
        ref=f"A1:{get_column_letter(len(headers))}{1 + len(data['ranged'])}",
    )
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium3", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)


def write_channel_breakdown(wb, data):
    ws = add_ws(wb, "Channel Breakdown")
    set_col_widths(ws, {"A": 30, "B": 12, "C": 12, "D": 12})

    section_title(ws, 1, "Weekly Metrics — Customer Raised Tickets", 4)
    write_header_row(ws, 2, ["Channel", "Open", "Closed", "Total"])

    rows_cust = [
        ("AI Bot",                 data["ch"]["AI Bot"]),
        ("Messaging / Live Chat",  data["ch"]["Messaging/Chat"]),
        ("Architect — Customer",   data["arch_b"]["customer"]),
    ]
    r = 3
    tot = [0, 0, 0]
    for i, (label, ch) in enumerate(rows_cust):
        o, c, t = len(ch["open"]), len(ch["closed"]), len(ch["tickets"])
        tot[0] += o; tot[1] += c; tot[2] += t
        bg = WHITE if i % 2 == 0 else ROW_ALT
        write_data_row(ws, r, [label, o, c, t], bg=bg)
        r += 1
    write_total_row(ws, r, ["TOTAL", *tot])
    r += 2

    section_title(ws, r, "Architect Breakdown", 4)
    r += 1
    write_header_row(ws, r, ["Sub-Bucket", "Open", "Closed", "Total"])
    r += 1
    arch_rows = [
        ("Customer tickets",        data["arch_b"]["customer"]),
        ("Internal — on behalf of", data["arch_b"]["internal"]),
        ("Failure notifications",   data["arch_b"]["failure_notification"]),
    ]
    at = [0, 0, 0]
    for i, (label, ch) in enumerate(arch_rows):
        o, c, t = len(ch["open"]), len(ch["closed"]), len(ch["tickets"])
        at[0] += o; at[1] += c; at[2] += t
        bg = WHITE if i % 2 == 0 else ROW_ALT
        write_data_row(ws, r, [label, o, c, t], bg=bg)
        r += 1
    write_total_row(ws, r, ["TOTAL", *at])


def write_category_performance(wb, data):
    ws = add_ws(wb, "Category Performance")
    set_col_widths(ws, {"A": 42, "B": 13, "C": 14, "D": 14, "E": 10})

    section_title(ws, 1, "Support Performance by Category — Last 4 Complete Weeks", 5)
    write_header_row(ws, 2, ["Category", "Ticket Count", "Median (hrs)", "Average (hrs)", "P90 (hrs)"])

    for i, cat in enumerate(data["cat_perf"]):
        bg = WHITE if i % 2 == 0 else ROW_ALT
        write_data_row(ws, 3 + i, [
            cat["label"], cat["count"],
            cat["median"] if cat["median"] is not None else "—",
            cat["avg"]    if cat["avg"]    is not None else "—",
            cat["p90"]    if cat["p90"]    is not None else "—",
        ], bg=bg)

    # Bar chart — ticket count by category
    if data["cat_perf"]:
        n = len(data["cat_perf"])
        chart = BarChart()
        chart.type    = "bar"
        chart.title   = "Tickets by Category (last 4 weeks)"
        chart.y_axis.title = "Count"
        chart.grouping = "clustered"
        chart.width   = 22
        chart.height  = 14

        cats_ref   = Reference(ws, min_col=1, min_row=3, max_row=3 + n - 1)
        counts_ref = Reference(ws, min_col=2, min_row=2, max_row=3 + n - 1)
        chart.add_data(counts_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = ORANGE
        ws.add_chart(chart, f"G2")


def write_week_trend(wb, data):
    ws = add_ws(wb, "Week Trend")
    set_col_widths(ws, {"A": 10, "B": 18, "C": 10})

    section_title(ws, 1, "Week-by-Week Ticket Volume (Year to Date)", 3)
    write_header_row(ws, 2, ["Week", "Date Range", "Ticket Count"])

    for i, w in enumerate(data["all_weeks"]):
        bg = WHITE if i % 2 == 0 else ROW_ALT
        write_data_row(ws, 3 + i, [w["label"], w["display"], w["count"]], bg=bg)

    # Line/bar chart
    n = len(data["all_weeks"])
    if n > 0:
        chart = BarChart()
        chart.title      = "Tickets per Week"
        chart.y_axis.title = "Tickets"
        chart.width      = 26
        chart.height     = 14

        labels = Reference(ws, min_col=1, min_row=3, max_row=3 + n - 1)
        counts = Reference(ws, min_col=3, min_row=2, max_row=3 + n - 1)
        chart.add_data(counts, titles_from_data=True)
        chart.set_categories(labels)
        chart.series[0].graphicalProperties.solidFill = ORANGE
        ws.add_chart(chart, "E2")


def write_top_issues(wb, data):
    ws = add_ws(wb, "Top Issues")

    weeks = data["recent_weeks"]
    cols  = ["Issue"] + [w["label"] for w in weeks] + ["Total"]
    set_col_widths(ws, {get_column_letter(i + 1): (34 if i == 0 else 10)
                        for i in range(len(cols))})

    section_title(ws, 1, "Top Issues — Last 4 Weeks", len(cols))
    write_header_row(ws, 2, cols)

    for i, (label, rx) in enumerate(TOP_ISSUES):
        counts = [len([t for t in w["tickets"]
                       if rx.search(t.get("subject", "") or "")])
                  for w in weeks]
        bg = WHITE if i % 2 == 0 else ROW_ALT
        write_data_row(ws, 3 + i, [label, *counts, sum(counts)], bg=bg)


def write_customer_volume(wb, data):
    ws = add_ws(wb, "Customer Volume")
    set_col_widths(ws, {"A": 28, "B": 10, "C": 10, "D": 10})

    section_title(ws, 1, "Top Customers by Volume (Period)", 4)
    write_header_row(ws, 2, ["Customer", "Open", "Solved", "Total"])

    sorted_custs = sorted(
        data["cust_map"].items(),
        key=lambda x: -(x[1]["open"] + x[1]["solved"])
    )[:20]

    for i, (cust, cnts) in enumerate(sorted_custs):
        o = cnts["open"]; s = cnts["solved"]; t = o + s
        bg = WHITE if i % 2 == 0 else ROW_ALT
        write_data_row(ws, 3 + i, [cust, o, s, t], bg=bg)

    # Horizontal bar chart
    n = len(sorted_custs)
    if n > 0:
        chart = BarChart()
        chart.type       = "bar"
        chart.title      = "Top Customers by Volume"
        chart.y_axis.title = "Tickets"
        chart.width      = 20
        chart.height     = max(12, n * 0.6)

        cats = Reference(ws, min_col=1, min_row=3, max_row=3 + n - 1)
        tots = Reference(ws, min_col=4, min_row=2, max_row=3 + n - 1)
        chart.add_data(tots, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = ORANGE
        ws.add_chart(chart, "F2")


def write_daily_volume(wb, data):
    ws = add_ws(wb, "Daily Volume")
    set_col_widths(ws, {"A": 14, "B": 8, "C": 10})

    section_title(ws, 1, "Daily Ticket Volume (Period)", 3)
    write_header_row(ws, 2, ["Date", "Day", "Ticket Count"])

    sorted_dates = sorted(data["daily"].keys())
    for i, d in enumerate(sorted_dates):
        try:
            weekday = datetime.strptime(d, "%Y-%m-%d").strftime("%A")
        except Exception:
            weekday = ""
        bg = EXCL_BG if datetime.strptime(d, "%Y-%m-%d").weekday() >= 5 else (WHITE if i % 2 == 0 else ROW_ALT)
        write_data_row(ws, 3 + i, [d, weekday, data["daily"][d]], bg=bg)

    n = len(sorted_dates)
    if n > 0:
        chart = BarChart()
        chart.title        = "Daily Volume"
        chart.y_axis.title = "Tickets"
        chart.width        = max(20, n * 0.4)
        chart.height       = 12

        cats   = Reference(ws, min_col=1, min_row=3, max_row=3 + n - 1)
        counts = Reference(ws, min_col=3, min_row=2, max_row=3 + n - 1)
        chart.add_data(counts, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = ORANGE
        ws.add_chart(chart, "E2")


def write_status_dist(wb, data):
    ws = add_ws(wb, "Status Distribution")
    set_col_widths(ws, {"A": 14, "B": 12})

    section_title(ws, 1, "Status Distribution (Period)", 2)
    write_header_row(ws, 2, ["Status", "Count"])

    sc = data["status_counts"]
    order = ["Open", "Pending", "Solved", "Closed"]
    statuses = order + [s for s in sc if s not in order]

    for i, status in enumerate(statuses):
        if sc.get(status):
            bg = WHITE if i % 2 == 0 else ROW_ALT
            write_data_row(ws, 3 + i, [status, sc[status]], bg=bg)

    total_r = 3 + len(statuses)
    write_total_row(ws, total_r, ["TOTAL", sum(sc.values())])

    # Donut chart
    STATUS_COLORS = {
        "Open": "E8612C", "Pending": "F4A261",
        "Solved": "2A9D8F", "Closed": "457B9D"
    }
    valid = [(s, sc[s]) for s in statuses if sc.get(s)]
    if valid:
        chart = PieChart()
        chart.title  = "Status Distribution"
        chart.width  = 16
        chart.height = 14

        cats   = Reference(ws, min_col=1, min_row=3, max_row=3 + len(valid) - 1)
        counts = Reference(ws, min_col=2, min_row=2, max_row=3 + len(valid) - 1)
        chart.add_data(counts, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "D2")


def write_resolution_time(wb, data):
    ws = add_ws(wb, "Resolution Time")
    set_col_widths(ws, {"A": 42, "B": 13, "C": 14, "D": 14, "E": 10})

    section_title(ws, 1, "Resolution Time by Category (Period)", 5)
    write_header_row(ws, 2, ["Category", "Ticket Count", "Median (hrs)", "Average (hrs)", "P90 (hrs)"])

    cats_with_data = [c for c in data["cat_range_stats"] if c["count"] > 0]
    for i, cat in enumerate(cats_with_data):
        bg = WHITE if i % 2 == 0 else ROW_ALT
        write_data_row(ws, 3 + i, [
            cat["label"], cat["count"],
            cat["median"] if cat["median"] is not None else "—",
            cat["avg"]    if cat["avg"]    is not None else "—",
            cat["p90"]    if cat["p90"]    is not None else "—",
        ], bg=bg)

    # Grouped bar chart — median vs average
    cats_with_times = [c for c in cats_with_data if c["median"] is not None]
    if cats_with_times:
        n     = len(cats_with_times)
        chart = BarChart()
        chart.title        = "Resolution Time by Category (hours)"
        chart.y_axis.title = "Hours"
        chart.grouping     = "clustered"
        chart.width        = 24
        chart.height       = 14

        cats_ref = Reference(ws, min_col=1, min_row=3,
                             max_row=3 + len(cats_with_data) - 1)
        med_ref  = Reference(ws, min_col=3, min_row=2,
                             max_row=3 + len(cats_with_data) - 1)
        avg_ref  = Reference(ws, min_col=4, min_row=2,
                             max_row=3 + len(cats_with_data) - 1)
        chart.add_data(med_ref, titles_from_data=True)
        chart.add_data(avg_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = ORANGE
        chart.series[1].graphicalProperties.solidFill = DARK
        ws.add_chart(chart, "G2")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Generate Klarity Support Excel report")
    parser.add_argument("--email",  required=True, help="Zendesk email")
    parser.add_argument("--token",  required=True, help="Zendesk API token")
    parser.add_argument("--weeks",  type=int, default=4,
                        help="Number of recent complete weeks to include (default: 4)")
    parser.add_argument("--output", default=None,
                        help="Output filename (default: klarity_report_YYYYMMDD.xlsx)")
    args = parser.parse_args()

    since  = since_n_weeks(args.weeks)
    output = args.output or f"klarity_report_{date.today().strftime('%Y%m%d')}.xlsx"

    print(f"\nKlarity Support — Excel Report Generator")
    print(f"  Subdomain : {SUBDOMAIN}")
    print(f"  Period    : {since} → {date.today().isoformat()} ({args.weeks} complete weeks)")
    print(f"  Output    : {output}\n")

    print("Fetching data from Zendesk…")
    try:
        data = load_all_data(args.email, args.token, since)
    except requests.exceptions.HTTPError as e:
        print(f"\nHTTP Error: {e}")
        if e.response is not None and e.response.status_code == 401:
            print("  → 401 Unauthorized — check your email and API token.")
        sys.exit(1)

    print(f"\nData loaded:")
    print(f"  Actionable tickets (all time) : {len(data['real'])}")
    print(f"  In selected period            : {len(data['ranged'])}")
    print(f"    Open                        : {len(data['open_t'])}")
    print(f"    Solved / Closed             : {len(data['closed_t'])}")
    print(f"  Unsolved in Group (live)      : {len(data['unsolved'])}")
    print(f"  Failed Ops excluded           : {data['fo_count']}")

    print("\nBuilding Excel workbook…")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    write_summary(wb, data, since, args.weeks)
    write_raw_data(wb, data)
    write_channel_breakdown(wb, data)
    write_category_performance(wb, data)
    write_week_trend(wb, data)
    write_top_issues(wb, data)
    write_customer_volume(wb, data)
    write_daily_volume(wb, data)
    write_status_dist(wb, data)
    write_resolution_time(wb, data)

    wb.save(output)
    print(f"\nReport saved: {output}")
    print("Sheets:")
    for ws in wb.worksheets:
        print(f"  • {ws.title}")


if __name__ == "__main__":
    main()
